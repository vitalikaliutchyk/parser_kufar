from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from urllib.parse import urljoin
from datetime import datetime, timedelta
import random
import json
import os
import asyncio
import requests
from bs4 import BeautifulSoup
from telegram import Bot
from telegram.error import TelegramError
from dotenv import load_dotenv

load_dotenv()

PAGES_URLS = [
    "https://www.kufar.by/l/kompyuternaya-tehnika/noutbuki/nb~apple?clp=v.or%3A66&cursor=eyJ0IjoiYWJzIiwiZiI6dHJ1ZSwicCI6MX0%3D&sort=lst.d",
    "https://www.kufar.by/l/kompyuternaya-tehnika/noutbuki/nb~apple?clp=v.or%3A66&cursor=eyJ0IjoiYWJzIiwiZiI6dHJ1ZWSJCI6MiwicGl0IjoiMjkxMTUxMjEifQ%3D%3D&sort=lst.d"
]

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36",
    "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7"
}
BASE_URL = "https://www.kufar.by/"
DATA_FILE = "data.json"
TELEGRAM_TOKEN = os.getenv("TELEGRAM_TOKEN")
CHAT_ID = os.getenv("CHAT_ID")


def format_price(price_str):
    try:
        return int(price_str.replace(" ", "").replace("р.", "").strip())
    except (ValueError, AttributeError):
        return None


def parse_datetime(time_element):
    if not time_element:
        return None

    try:
        time_str = time_element.get_text(strip=True)
        if "Сегодня" in time_str:
            time_str = time_str.replace("Сегодня, ", "")
            time_part = datetime.strptime(time_str, "%H:%M")
            return datetime.now().replace(hour=time_part.hour, minute=time_part.minute, second=0, microsecond=0)
        elif "Вчера" in time_str:
            time_str = time_str.replace("Вчера, ", "")
            time_part = datetime.strptime(time_str, "%H:%M")
            date = datetime.now() - timedelta(days=1)
            return date.replace(hour=time_part.hour, minute=time_part.minute, second=0, microsecond=0)
        else:
            return datetime.strptime(time_str, "%d.%m.%Y %H:%M")
    except Exception as e:
        print(f"Ошибка парсинга времени: {str(e)}")
        return None


def parse_page(url):
    try:
        response = requests.get(url, headers=HEADERS, timeout=15)
        response.raise_for_status()
    except requests.RequestException as e:
        print(f"Ошибка запроса {url}: {str(e)}")
        return []

    soup = BeautifulSoup(response.text, "html.parser")
    container = soup.find("div", {"class": "styles_cards__bBppJ"})

    if not container:
        print(f"Контейнер объявлений не найден на странице {url}!")
        return []

    listings = []
    for section in container.find_all("section"):
        item = section.find("a", {"data-testid": "kufar-ad"})
        if not item:
            continue

        try:
            title = item.find("h3", class_="styles_title__F3uIe").get_text(strip=True) if item.find("h3",
                                                                                                    class_="styles_title__F3uIe") else "Без названия"
            price_element = item.find("p", class_="styles_price__aVxZc")
            price = format_price(price_element.get_text(strip=True)) if price_element else None
            region_element = item.find("p", class_="styles_region__qCRbf")
            region = region_element.get_text(strip=True).replace("Минск, ", "") if region_element else "Не указан"
            time_element = item.find("span", class_="styles_secondary__MzdEb")
            parsed_time = parse_datetime(time_element)
            raw_href = item.get("href", "")
            clean_href = raw_href.split('?')[0]
            link = urljoin(BASE_URL, clean_href)

            listings.append({
                "title": title,
                "price": price,
                "region": region,
                "time": parsed_time.strftime("%d.%m.%Y %H:%M") if parsed_time else None,
                "link": link
            })

        except Exception as e:
            print(f"Ошибка парсинга объявления: {str(e)}")
            continue

    return listings


def create_excel_file(data, filename="noutbuki.xlsx"):
    if not data:
        print("Нет данных для создания файла Excel")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Ноутбуки Apple"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    headers = ["№", "Название", "Цена (BYN)", "Регион", "Дата публикации", "Ссылка"]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for idx, item in enumerate(data, 1):
        ws.append([
            idx,
            item.get("title", "Нет данных"),
            item.get("price", "Нет данных"),
            item.get("region", "Нет данных"),
            item.get("time", "Нет данных"),
            item.get("link", "Нет данных")
        ])

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=idx + 1, column=col)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    column_widths = {"A": 8, "B": 50, "C": 15, "D": 20, "E": 20, "F": 60}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=3, max_col=3):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0" р."'

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
        for cell in row:
            cell.style = "Hyperlink"

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    try:
        wb.save(filename)
        print(f"\nФайл {filename} успешно создан!")
    except PermissionError:
        print(f"\nОшибка: Не удалось сохранить файл {filename}. Возможно, файл открыт в другой программе.")


def load_data():
    if not os.path.exists(DATA_FILE):
        return []

    try:
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, UnicodeDecodeError) as e:
        print(f"Ошибка чтения файла данных: {str(e)}")
        return []


def save_data(data):
    try:
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        print("Данные успешно сохранены.")
    except IOError as e:
        print(f"Ошибка сохранения данных: {str(e)}")


def find_changes(old_data, new_data):
    old_dict = {item['link']: item for item in old_data}
    new_dict = {item['link']: item for item in new_data}

    new_items = [item for link, item in new_dict.items() if link not in old_dict]
    updated_items = []

    for link, item in new_dict.items():
        if link in old_dict:
            old_item = old_dict[link]
            if item['price'] != old_item['price']:
                updated_items.append(item)
            elif item['title'] != old_item['title'] or item['region'] != old_item['region']:
                updated_items.append(item)

    print(f"Найдено новых объявлений: {len(new_items)}, обновлённых: {len(updated_items)}")
    return new_items, updated_items


def format_message(item, is_new):
    emoji = "🆕" if is_new else "🔄"
    price = f"{item['price']} BYN" if item.get('price') else "Цена не указана"
    return (
        f"{emoji} {'Новое объявление' if is_new else 'Обновление'}\n"
        f"📌 {item.get('title', 'Без названия')}\n"
        f"💰 {price}\n"
        f"📍 {item.get('region', 'Регион не указан')}\n"
        f"🕒 {item.get('time', 'Время не указано')}\n"
        f"🔗 {item.get('link', 'Ссылка отсутствует')}"
    )


async def send_telegram_notification(bot, new_items, updated_items):
    if not TELEGRAM_TOKEN or not CHAT_ID:
        print("Токен Telegram или ID чата не установлены. Уведомления не будут отправлены.")
        return

    if not new_items and not updated_items:
        print("Нет новых или обновлённых объявлений для отправки.")
        return

    total_sent = 0
    for item in new_items:
        message = format_message(item, is_new=True)
        try:
            await bot.send_message(chat_id=CHAT_ID, text=message, disable_web_page_preview=True)
            await asyncio.sleep(random.uniform(1, 2))  # Добавляем случайную задержку
            total_sent += 1
        except TelegramError as e:
            print(f"Ошибка отправки сообщения: {e}")

    for item in updated_items:
        message = format_message(item, is_new=False)
        try:
            await bot.send_message(chat_id=CHAT_ID, text=message, disable_web_page_preview=True)
            await asyncio.sleep(random.uniform(1, 2))
            total_sent += 1
        except TelegramError as e:
            print(f"Ошибка отправки сообщения: {e}")

    print(f"Успешно отправлено сообщений: {total_sent}")


async def job(bot):
    print(f"\nНачало парсинга в {datetime.now().strftime('%d.%m.%Y %H:%M:%S')}...")
    all_items = []
    for url in PAGES_URLS:
        items = parse_page(url)
        all_items.extend(items)
        print(f"Спарсено {len(items)} объявлений с {url}")
        await asyncio.sleep(random.uniform(2, 4))

    if all_items:
        create_excel_file(all_items)
        old_data = load_data()
        new_items, updated_items = find_changes(old_data, all_items)

        if new_items or updated_items:
            await send_telegram_notification(bot, new_items, updated_items)
            save_data(all_items)
        else:
            print("Изменений нет, данные не сохранены.")
    else:
        print("Не удалось получить данные с сайта.")


async def main():
    if not TELEGRAM_TOKEN or not CHAT_ID:
        print("Предупреждение: Токен Telegram или ID чата не установлены. Уведомления не будут отправляться.")

    bot = Bot(token=TELEGRAM_TOKEN) if TELEGRAM_TOKEN else None

    while True:
        try:
            await job(bot)
            await asyncio.sleep(3 * 3600)  # Ожидание 3 часа перед следующим запуском
        except Exception as e:
            print(f"Критическая ошибка в основном цикле: {str(e)}")
            await asyncio.sleep(3600)  # Ожидание 1 час после ошибки


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        print("\nРабота программы остановлена пользователем")
    except Exception as e:
        print(f"Критическая ошибка: {str(e)}")
